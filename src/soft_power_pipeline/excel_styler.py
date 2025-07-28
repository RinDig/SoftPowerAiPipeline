"""Excel styling and formatting utilities"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """Handles Excel worksheet styling and formatting"""
    
    def __init__(self):
        # Define color palette
        self.colors = {
            # Headers
            'header_bg': 'D5E8D4',        # Light green
            'header_text': '2D5F2B',      # Dark green
            
            # Categories
            'infrastructure_bg': 'FFE6CC',  # Light orange
            'assets_bg': 'E1D5E7',         # Light purple  
            'outcome_bg': 'D4E6F1',        # Light blue
            
            # Scores (traffic light system)
            'high_score': '90EE90',        # Light green
            'medium_score': 'FFD700',      # Gold
            'low_score': 'FFB6C1',         # Light pink
            
            # Text
            'dark_text': '2C3E50',
            'medium_text': '5D6D7E',
            'light_text': '85929E',
            
            # Borders
            'border_color': 'BDC3C7'
        }
        
        # Define fonts
        self.fonts = {
            'header': Font(name='Calibri', size=12, bold=True, color=self.colors['header_text']),
            'subheader': Font(name='Calibri', size=11, bold=True, color=self.colors['dark_text']),
            'body': Font(name='Calibri', size=10, color=self.colors['dark_text']),
            'small': Font(name='Calibri', size=9, color=self.colors['medium_text'])
        }
        
        # Define fills
        self.fills = {
            'header': PatternFill(start_color=self.colors['header_bg'], 
                                 end_color=self.colors['header_bg'], 
                                 fill_type='solid'),
            'infrastructure': PatternFill(start_color=self.colors['infrastructure_bg'], 
                                        end_color=self.colors['infrastructure_bg'], 
                                        fill_type='solid'),
            'assets': PatternFill(start_color=self.colors['assets_bg'], 
                                end_color=self.colors['assets_bg'], 
                                fill_type='solid'),
            'outcome': PatternFill(start_color=self.colors['outcome_bg'], 
                                 end_color=self.colors['outcome_bg'], 
                                 fill_type='solid'),
            'high_score': PatternFill(start_color=self.colors['high_score'], 
                                    end_color=self.colors['high_score'], 
                                    fill_type='solid'),
            'medium_score': PatternFill(start_color=self.colors['medium_score'], 
                                      end_color=self.colors['medium_score'], 
                                      fill_type='solid'),
            'low_score': PatternFill(start_color=self.colors['low_score'], 
                                   end_color=self.colors['low_score'], 
                                   fill_type='solid')
        }
        
        # Define borders
        thin_border = Side(border_style="thin", color=self.colors['border_color'])
        self.borders = {
            'thin': Border(left=thin_border, right=thin_border, 
                          top=thin_border, bottom=thin_border),
            'thick_bottom': Border(bottom=Side(border_style="medium", 
                                             color=self.colors['border_color']))
        }
        
        # Define alignment
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
            'wrap': Alignment(horizontal='left', vertical='top', wrap_text=True)
        }
    
    def style_header_row(self, ws, row_num, max_col):
        """Style header row with bold font and background color"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.border = self.borders['thin']
            cell.alignment = self.alignments['center']
    
    def style_data_rows(self, ws, start_row, end_row, category_col=None):
        """Style data rows with alternating colors and category-based coloring"""
        max_col = ws.max_column
        
        for row in range(start_row, end_row + 1):
            # Determine category for row coloring
            category = None
            if category_col:
                category_cell = ws.cell(row=row, column=category_col)
                if category_cell.value:
                    category = str(category_cell.value).lower()
            
            # Apply row styling
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.fonts['body']
                cell.border = self.borders['thin']
                cell.alignment = self.alignments['left']
                
                # Apply category-specific background color
                if category_col and col == category_col:
                    if 'infrastructure' in category:
                        cell.fill = self.fills['infrastructure']
                    elif 'asset' in category:
                        cell.fill = self.fills['assets']
                    elif 'outcome' in category:
                        cell.fill = self.fills['outcome']
    
    def style_score_columns(self, ws, score_columns, start_row, end_row):
        """Apply conditional formatting to score columns"""
        for col_letter in score_columns:
            col_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            
            # Apply color scale: red (low) -> yellow (medium) -> green (high)
            color_scale = ColorScaleRule(
                start_type='min', start_color='FFB6C1',  # Light pink
                mid_type='percentile', mid_value=50, mid_color='FFD700',  # Gold
                end_type='max', end_color='90EE90'  # Light green
            )
            ws.conditional_formatting.add(col_range, color_scale)
    
    def style_impact_scores(self, ws, impact_col, start_row, end_row):
        """Special styling for impact scores (0-10 scale)"""
        col_letter = get_column_letter(impact_col)
        
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=impact_col)
            if cell.value and isinstance(cell.value, (int, float)):
                score = float(cell.value)
                
                # Color code based on impact score
                if score >= 7:  # High impact
                    cell.fill = self.fills['high_score']
                    cell.font = Font(name='Calibri', size=10, bold=True, color='2D5F2B')
                elif score >= 4:  # Medium impact
                    cell.fill = self.fills['medium_score']
                    cell.font = Font(name='Calibri', size=10, bold=True, color='B7950B')
                else:  # Low impact
                    cell.fill = self.fills['low_score']
                    cell.font = Font(name='Calibri', size=10, color='A93226')
    
    def add_section_headers(self, ws, row, col, text):
        """Add styled section headers"""
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = self.fonts['subheader']
        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        cell.border = self.borders['thick_bottom']
        cell.alignment = self.alignments['left']
        
        # Merge cells if it's a section header
        if col == 1:
            try:
                ws.merge_cells(f"A{row}:E{row}")
            except:
                pass  # If merge fails, continue
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths with reasonable limits"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else None
            
            if not column_letter:
                continue
                
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width based on column type
            if column_letter == 'A':  # Country
                adjusted_width = min(max_length + 2, 15)
            elif column_letter == 'B':  # Data Point/Item
                adjusted_width = min(max_length + 2, 35)
            elif column_letter == 'C':  # Raw Value
                adjusted_width = min(max_length + 2, 25)
            elif column_letter == 'D':  # Category
                adjusted_width = min(max_length + 2, 15)
            elif 'Score' in str(ws.cell(1, column[0].column).value or ''):  # Score columns
                adjusted_width = min(max_length + 2, 12)
            else:  # Other columns
                adjusted_width = min(max_length + 2, 40)
            
            ws.column_dimensions[column_letter].width = max(adjusted_width, 8)
    
    def freeze_header_row(self, ws):
        """Freeze the top row for easier scrolling"""
        ws.freeze_panes = ws['A2']
    
    def style_correlation_table(self, ws, start_row, end_row):
        """Special styling for correlation tables"""
        max_col = ws.max_column
        
        # Style header row
        self.style_header_row(ws, start_row, max_col)
        
        # Style data rows
        for row in range(start_row + 1, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.fonts['body']
                cell.border = self.borders['thin']
                cell.alignment = self.alignments['center']
                
                # Highlight correlation values
                if col > 1 and cell.value and isinstance(cell.value, (int, float)):
                    if abs(float(cell.value)) > 0.7:  # Strong correlation
                        cell.font = Font(name='Calibri', size=10, bold=True, color='2D5F2B')
                    elif abs(float(cell.value)) > 0.4:  # Medium correlation
                        cell.font = Font(name='Calibri', size=10, bold=True, color='B7950B')
    
    def add_legend(self, ws, start_row):
        """Add a color legend for categories"""
        legend_items = [
            ("Infrastructure", self.fills['infrastructure']),
            ("Assets", self.fills['assets']),
            ("Outcomes", self.fills['outcome'])
        ]
        
        ws.cell(row=start_row, column=1, value="Color Legend:").font = self.fonts['subheader']
        
        for i, (category, fill) in enumerate(legend_items, 1):
            cell = ws.cell(row=start_row + i, column=1, value=f"  {category}")
            cell.font = self.fonts['body']
            cell.fill = fill
            cell.border = self.borders['thin']