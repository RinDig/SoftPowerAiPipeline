"""Excel data extraction module with block-based parsing"""

import logging
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
import pandas as pd
from openpyxl import load_workbook

from ..models.excel_models import (
    DataItem,
    DataBlock,
    NationalLevelReport,
    OrganizationProfile,
    CountryWorkbook
)


logger = logging.getLogger(__name__)


class ExcelExtractor:
    """Extract structured data from standardized country Excel workbooks"""
    
    # Column mapping based on PRD specification
    COLUMN_MAP = {
        'A': 'category',
        'B': 'item',
        'C': 'value',
        'D': 'notes',
        'E': 'source'
    }
    
    # Keywords to identify sheet types
    NATIONAL_KEYWORDS = ['national level', 'national-level', 'country level']
    SKIP_SHEETS = ['template', 'instructions', 'readme']
    
    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []
    
    def extract_country_data(self, excel_path: Union[str, Path]) -> Optional[CountryWorkbook]:
        """Extract all data from a country Excel workbook"""
        excel_path = Path(excel_path)
        
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            return None
        
        try:
            # Extract country name from filename
            country = self._extract_country_name(excel_path.name)
            
            # Load workbook
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            
            # Initialize country workbook
            country_wb = CountryWorkbook(country=country)
            
            # Process each sheet
            for sheet_name in sheet_names:
                if self._should_skip_sheet(sheet_name):
                    continue
                
                logger.info(f"Processing sheet: {sheet_name}")
                
                if self._is_national_level_sheet(sheet_name):
                    # Process as national level sheet
                    national_report = self._extract_national_level_data(
                        workbook[sheet_name], country
                    )
                    if national_report:
                        country_wb.national_report = national_report
                else:
                    # Process as organization sheet
                    org_profile = self._extract_organization_data(
                        workbook[sheet_name], country, sheet_name
                    )
                    if org_profile:
                        country_wb.organization_profiles.append(org_profile)
            
            workbook.close()
            
            # Log any warnings or errors
            if self.warnings:
                logger.warning(f"Warnings for {country}: {self.warnings}")
            if self.errors:
                logger.error(f"Errors for {country}: {self.errors}")
            
            return country_wb
            
        except Exception as e:
            logger.error(f"Error processing {excel_path}: {str(e)}")
            return None
    
    def _extract_country_name(self, filename: str) -> str:
        """Extract country name from filename"""
        # Remove date prefix and extension
        # Format: "240312 Australia.xlsx" -> "Australia"
        parts = filename.split()
        if len(parts) >= 2:
            # Remove extension from last part
            country = ' '.join(parts[1:]).replace('.xlsx', '').replace('.xls', '')
            return country.strip()
        return filename.replace('.xlsx', '').replace('.xls', '').strip()
    
    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """Check if sheet should be skipped"""
        sheet_lower = sheet_name.lower()
        return any(skip in sheet_lower for skip in self.SKIP_SHEETS)
    
    def _is_national_level_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is national level data"""
        sheet_lower = sheet_name.lower()
        return any(keyword in sheet_lower for keyword in self.NATIONAL_KEYWORDS)
    
    def _extract_national_level_data(self, sheet, country: str) -> Optional[NationalLevelReport]:
        """Extract data from national level sheet"""
        try:
            # Convert sheet to DataFrame
            df = self._sheet_to_dataframe(sheet)
            
            if df.empty:
                self.warnings.append(f"Empty national level sheet for {country}")
                return None
            
            # Extract blocks
            data_blocks = self._extract_data_blocks(df)
            
            return NationalLevelReport(
                country=country,
                data_blocks=data_blocks
            )
            
        except Exception as e:
            self.errors.append(f"Error processing national level sheet: {str(e)}")
            return None
    
    def _extract_organization_data(self, sheet, country: str, sheet_name: str) -> Optional[OrganizationProfile]:
        """Extract data from organization sheet"""
        try:
            # Convert sheet to DataFrame
            df = self._sheet_to_dataframe(sheet)
            
            if df.empty:
                self.warnings.append(f"Empty sheet: {sheet_name}")
                return None
            
            # Extract blocks
            data_blocks = self._extract_data_blocks(df)
            
            # Create organization profile
            org_profile = OrganizationProfile(
                country=country,
                organization_name=sheet_name,
                data_blocks=data_blocks
            )
            
            # Try to extract key fields from blocks
            self._populate_organization_fields(org_profile)
            
            return org_profile
            
        except Exception as e:
            self.errors.append(f"Error processing organization sheet {sheet_name}: {str(e)}")
            return None
    
    def _sheet_to_dataframe(self, sheet) -> pd.DataFrame:
        """Convert Excel sheet to pandas DataFrame"""
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            data.append(row)
        
        if not data:
            return pd.DataFrame()
        
        # Create DataFrame with column names based on COLUMN_MAP
        max_cols = max(len(row) for row in data)
        columns = []
        for i in range(max_cols):
            col_letter = chr(ord('A') + i)
            columns.append(self.COLUMN_MAP.get(col_letter, f'col_{col_letter}'))
        
        df = pd.DataFrame(data, columns=columns[:len(data[0])])
        return df
    
    def _extract_data_blocks(self, df: pd.DataFrame) -> List[DataBlock]:
        """Extract data blocks from DataFrame using block-based parsing"""
        blocks = []
        current_category = None
        current_items = []
        
        for idx, row in df.iterrows():
            # Check if this is a new category
            if pd.notna(row.get('category')):
                # Save previous block if exists
                if current_category and current_items:
                    blocks.append(DataBlock(
                        category=current_category,
                        data_points=current_items
                    ))
                
                # Start new block
                current_category = str(row['category']).strip()
                current_items = []
            
            # Add item to current block if we have item and value
            if pd.notna(row.get('item')) and pd.notna(row.get('value')):
                try:
                    item = DataItem(
                        item=str(row['item']).strip(),
                        value=self._parse_value(row['value']),
                        notes=str(row['notes']).strip() if pd.notna(row.get('notes')) else None,
                        source=str(row['source']).strip() if pd.notna(row.get('source')) else None
                    )
                    current_items.append(item)
                except Exception as e:
                    self.warnings.append(f"Error parsing row {idx}: {str(e)}")
        
        # Don't forget the last block
        if current_category and current_items:
            blocks.append(DataBlock(
                category=current_category,
                data_points=current_items
            ))
        
        return blocks
    
    def _parse_value(self, value: Any) -> Union[str, int, float]:
        """Parse value to appropriate type"""
        if pd.isna(value):
            return ""
        
        # Try to convert to number if possible
        if isinstance(value, (int, float)):
            return value
        
        # Try to parse string as number
        value_str = str(value).strip()
        
        # Check for percentage
        if value_str.endswith('%'):
            try:
                return float(value_str.rstrip('%')) / 100
            except ValueError:
                pass
        
        # Try integer
        try:
            return int(value_str)
        except ValueError:
            pass
        
        # Try float
        try:
            return float(value_str)
        except ValueError:
            pass
        
        # Return as string
        return value_str
    
    def _populate_organization_fields(self, org_profile: OrganizationProfile) -> None:
        """Extract key organization fields from data blocks"""
        # Try to find lead ministry in accountability block
        accountability = org_profile.get_accountability_info()
        if accountability:
            lead_ministry = accountability.get_item_value("Lead Ministry")
            if lead_ministry:
                org_profile.lead_ministry = str(lead_ministry)
            
            mission = accountability.get_item_value("Mission")
            if mission:
                org_profile.mission = str(mission)
        
        # Try to find budget in finance block
        finance = org_profile.get_finance_info()
        if finance:
            budget = finance.get_item_value("Annual Budget")
            if not budget:
                budget = finance.get_item_value("Budget")
            if budget:
                org_profile.budget = str(budget)